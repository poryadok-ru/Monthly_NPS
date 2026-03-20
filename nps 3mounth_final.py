import os
import pandas as pd
import numpy as np
import time
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
import math

def split_dataframe_to_sheets(df, max_rows=1000000):
    sheets_data = {}
    total_parts = math.ceil(len(df) / max_rows)
    for part in range(total_parts):
        start_idx = part * max_rows
        end_idx = min((part + 1) * max_rows, len(df))
        sheet_name = f"общее_часть{part + 1}" if part > 0 else "общее"
        sheets_data[sheet_name] = df.iloc[start_idx:end_idx].copy()
    return sheets_data

def get_month_name(month_number):
    months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
              "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    try:
        return months[int(month_number) - 1] if not pd.isnull(month_number) else np.nan
    except Exception:
        return np.nan

def add_totals_row(pivot, base_cols, months_order):
    good_cols = [f"Хорошая {m}" for m in months_order if f"Хорошая {m}" in pivot.columns]
    bad_cols = [f"Плохая {m}" for m in months_order if f"Плохая {m}" in pivot.columns]
    neut_cols = [f"Нейтральная {m}" for m in months_order if f"Нейтральная {m}" in pivot.columns]
    cnt_cols = []
    for met in ["Количество отзывов", "Всего"]:
        cnt_cols.extend([f"{met} {m}" for m in months_order if f"{met} {m}" in pivot.columns])
    nps_cols = [f"NPS {m}" for m in months_order if f"NPS {m}" in pivot.columns]
    itog_row = {}

    for col in good_cols + bad_cols + neut_cols + cnt_cols:
        itog_row[col] = pivot[col].sum()
    for nps_col, good_col, bad_col, cnt_col in zip(nps_cols, good_cols, bad_cols, cnt_cols):
        good_sum = pivot[good_col].sum()
        bad_sum = pivot[bad_col].sum()
        cnt_sum = pivot[cnt_col].sum()
        itog_row[nps_col] = (good_sum - bad_sum) / cnt_sum if cnt_sum > 0 else np.nan
    if 'Количество отзывов Общий итог' in pivot.columns:
        itog_row['Количество отзывов Общий итог'] = pivot['Количество отзывов Общий итог'].sum()
    if 'NPS Общий итог' in pivot.columns and len(good_cols) > 0 and len(bad_cols) > 0:
        sum_good = pivot[good_cols].sum().sum()
        sum_bad = pivot[bad_cols].sum().sum()
        sum_total = 0
        for cnt_col in cnt_cols:
            sum_total += pivot[cnt_col].sum()
        itog_row['NPS Общий итог'] = (sum_good - sum_bad) / sum_total if sum_total > 0 else np.nan
    for bc in base_cols:
        itog_row[bc] = 'ИТОГО'
    for col in pivot.columns:
        if col not in itog_row:
            itog_row[col] = np.nan
    # Обнуляем СТМ в итоговой строке (по желанию)
    if 'СТМ' in pivot.columns:
        itog_row['СТМ'] = np.nan
    return pd.concat([pivot, pd.DataFrame([itog_row])], ignore_index=True)

def npsview_one(excel_file, add_message):
    t = time.time()
    add_message('\nНачинаю группировку комментариев!')
    today = datetime.now()
    last_month = today.replace(day=1) - timedelta(days=1)
    try:
        xl_file = pd.ExcelFile(excel_file)
        sheet_names = xl_file.sheet_names
        general_sheets = [name for name in sheet_names if name.startswith('общее')]
        products_sheet = next((name for name in sheet_names if name == 'по кодам'), None)
        if not general_sheets or not products_sheet:
            raise KeyError("Не найдены необходимые листы")
        dfs_general = [pd.read_excel(excel_file, sheet_name=sheet) for sheet in general_sheets]
        df_comments = pd.concat(dfs_general, ignore_index=True)
        df_products = pd.read_excel(excel_file, sheet_name=products_sheet)
        # Проверяем наличие СТМ
        if 'СТМ' not in df_products.columns:
            df_products['СТМ'] = np.nan
        stm_idx = df_products.columns.get_loc("СТМ")

        df_comments = df_comments[df_comments["Комментарий"].notna()]
        df_comments = df_comments[df_comments["Комментарий"].str.strip() != "-"]

        def safe_to_datetime(x):
            try:
                if pd.isna(x):
                    return pd.NaT
                if isinstance(x, (int, float)):
                    return pd.to_datetime("1899-12-30") + pd.to_timedelta(int(x), "D")
                return pd.to_datetime(x, errors="coerce", dayfirst=True)
            except Exception:
                return pd.NaT

        df_comments["Месяц и год"] = df_comments["Месяц и год"].apply(safe_to_datetime)
        df_comments = df_comments[df_comments["Месяц и год"].notna()]
        code_col = None
        for c in ["Код продукта", "Код (доп.)"]:
            if c in df_comments.columns and c in df_products.columns:
                code_col = c
                break
        if code_col is None:
            code_col = df_products.columns[0]

        product_comments = df_comments.groupby(code_col).apply(
            lambda x: list(zip(x["Комментарий"], x["Месяц и год"]))
        ).to_dict()

        rows_to_write = []
        is_code_row = []
        is_yellow_row = []
        for idx, row in df_products.iterrows():
            rows_to_write.append(list(row) + [None])  # карточка товара+is_yellow
            is_code_row.append(True)
            is_yellow_row.append(False)
            code = row[code_col] if code_col in row else row.iloc[0]
            supplier = row.iloc[2] if len(row) > 2 else None
            manager = row.iloc[3] if len(row) > 3 else None
            has_yellow_comment = False
            comments_data = product_comments.get(code, [])
            comment_rows = []
            yellow_flags = []
            for comment, cdate in comments_data:
                is_yellow = (
                        pd.notnull(cdate) and
                        isinstance(cdate, pd.Timestamp) and
                        cdate.year == last_month.year and
                        cdate.month == last_month.month
                )
                comment_row = [None] * len(row)
                comment_row[1] = comment
                if len(row) > 2:
                    comment_row[2] = supplier
                if len(row) > 3:
                    comment_row[3] = manager
                # Вставить СТМ в правильное место
                comment_row[stm_idx] = row['СТМ'] if 'СТМ' in row else np.nan
                comment_row.append(is_yellow)
                comment_rows.append(comment_row)
                yellow_flags.append(is_yellow)
                if is_yellow:
                    has_yellow_comment = True
            rows_to_write.extend(comment_rows)
            is_code_row.extend([False] * len(comment_rows))
            is_yellow_row.extend(yellow_flags)
            if has_yellow_comment:
                is_yellow_row[-(len(comment_rows) + 1)] = True

        out_df = pd.DataFrame(rows_to_write)
        out_df.columns = list(df_products.columns) + ["_yellow"]

        output_file = os.path.splitext(excel_file)[0] + '_upd.xlsx'

        with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
            out_df.drop(columns="_yellow").to_excel(writer, sheet_name="по кодам", index=False)
            general_sheets_data = split_dataframe_to_sheets(df_comments)
            for sheet_name, sheet_data in general_sheets_data.items():
                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(output_file)
        if "по кодам" in wb.sheetnames:
            sh = wb["по кодам"]
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            wrap_align = Alignment(wrap_text=True, vertical='center')
            header_font = Font(bold=True)
            is_code_row2 = [0] + is_code_row
            is_yellow_row2 = [False] + is_yellow_row
            max_row = sh.max_row
            max_col = sh.max_column

            for col in range(1, max_col + 1):
                cell = sh.cell(row=1, column=col)
                cell.alignment = wrap_align
                cell.font = header_font
                header_value = cell.value
                if header_value:
                    header_str = str(header_value)
                    if "NPS" in header_str and "Общий" in header_str:
                        cell.fill = green_fill
                    elif "NPS" in header_str:
                        cell.fill = blue_fill

            nps_cols = [col for col in range(1, max_col + 1)
                        if sh.cell(row=1, column=col).value and str(sh.cell(row=1, column=col).value).startswith("NPS ")]
            for col in nps_cols:
                for row in range(2, max_row + 1):
                    sh.cell(row=row, column=col).number_format = '0%'

            for i in range(2, max_row + 1):
                if is_yellow_row2[i - 1]:
                    for col in range(1, max_col + 1):
                        sh.cell(row=i, column=col).fill = yellow_fill

            i = 2
            while i <= max_row:
                if is_code_row2[i - 1]:
                    group_start = i + 1
                    group_end = group_start
                    while group_end <= max_row and not is_code_row2[group_end - 1]:
                        group_end += 1
                    if group_end > group_start:
                        sh.row_dimensions.group(start=group_start, end=group_end - 1, hidden=False)
                    i = group_end
                else:
                    i += 1

            for col in range(1, max_col + 1):
                cell = sh.cell(row=max_row, column=col)
                cell.font = Font(bold=True)

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('общее'):
                sh = wb[sheet_name]
                max_col = sh.max_column
                for col in range(1, max_col + 1):
                    cell = sh.cell(row=1, column=col)
                    cell.alignment = wrap_align
                    cell.font = header_font
        wb.save(output_file)
        print(f"Комментарии добавлены и строки сгруппированы в файл '{output_file}'")
    except FileNotFoundError:
        print(f"Ошибка: Файл '{excel_file}' не найден.")
    except KeyError as e:
        print(f"Ошибка: Лист с именем '{e}' не найден в Excel файле.")
    except Exception as e:
        print(f"Произошла ошибка: {e}")
    add_message(f'Готово за {time.time() - t:.1f} сек.')

def main():
    timer0 = time.perf_counter()
    folder = 'Исходники WB и Ozon'
    card_path = os.path.join(folder, 'Карточка товара.xlsx')
    save_path = os.path.join(folder, 'WB&OZON NPS 3 месяца.xlsx')
    print('Старт. Проверка условий.')
    print('Читаю файлы...')
    t1 = time.perf_counter()
    files = [f for f in os.listdir(folder)
             if (f.endswith('.xlsx') or f.endswith('.csv')) and not ('nps' in f.lower() or 'карточк' in f.lower())]
    dataframes = []
    for f in files:
        path = os.path.join(folder, f)
        try:
            if f.endswith('.csv'):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str)
            dataframes.append(df)
            print(f'  Прочитан: {f} ({len(df)} строк)')
        except Exception as e:
            print(f'Ошибка в файле {f}: {e}')
    if len(dataframes) == 0:
        print('Нет исходных данных!')
        return
    df_all = pd.concat(dataframes, ignore_index=True)
    print(f'Считано файлов: {len(dataframes)}, итоговых строк: {df_all.shape[0]}')
    print(f'Время чтения файлов: {time.perf_counter() - t1:.2f} сек.')

    print('Собираю файл...')
    rename_map = {
        "COMMENT": "Комментарий",
        "CREATED_AT": "Месяц и год",
        "PRODUCT.CODE": "Код продукта",
        "RATING": "Оценка",
        "SOURCE": "Источник",
        "IS_ABOUT_DELIVERY": "Про доставку"
    }
    df_all = df_all.rename(columns={col: rename_map[col] for col in df_all.columns if col in rename_map})

    print('Преобразовываю даты...')
    df_all['date_clean'] = df_all['Месяц и год'].fillna('').str.split('@').str[0].str.strip()
    dt = pd.to_datetime(df_all['date_clean'], format='%b %d, %Y', errors='coerce')
    not_parsed = dt.isna()
    if not_parsed.any():
        dt2 = pd.to_datetime(df_all.loc[not_parsed, 'date_clean'], dayfirst=True, errors='coerce')
        dt.loc[not_parsed] = dt2
    df_all['Дата_dt'] = dt
    df_all['Месяц и год'] = df_all['Дата_dt'].dt.date
    df_all['Год'] = df_all['Дата_dt'].dt.year
    df_all['Месяц'] = df_all['Дата_dt'].dt.month

    print('Работаю с оценками...')
    df_all["Результат"] = (
        df_all["Оценка"].map({'5': "Хорошая", '4': "Нейтральная", '3': "Плохая", '2': "Плохая", '1': "Плохая"})
    )
    df_all["Хорошая"] = (df_all["Результат"] == "Хорошая").astype("Int8")
    df_all["Нейтральная"] = (df_all["Результат"] == "Нейтральная").astype("Int8")
    df_all["Плохая"] = (df_all["Результат"] == "Плохая").astype("Int8")

    print('Заполяняю данные из Карточки товаров...')
    t5 = time.perf_counter()
    extra_cols = ["Продукт", "Основной менеджер", "СегментСтелажногоХранения", "Подгруппа 1", "Поставщик", "СТМ"]
    if os.path.exists(card_path):
        card_df = pd.read_excel(card_path, dtype=str)
        card_df = card_df.rename(columns=lambda x: x.strip())
        for col in extra_cols:
            if col not in card_df.columns:
                card_df[col] = np.nan
        card_df.set_index("Код (доп.)", inplace=True)
        df_all = df_all.join(card_df[extra_cols], on="Код продукта", how="left")
    else:
        for c in extra_cols:
            df_all[c] = np.nan

    print('Играю в нарды...')
    final_headers = [
        "Комментарий", "Месяц и год", "Год", "Месяц", "Код продукта", "Продукт", "Основной менеджер",
        "Оценка", "Результат", "Хорошая", "Нейтральная", "Плохая", "Источник",
        "СегментСтелажногоХранения", "Подгруппа 1", "Поставщик", "СТМ", "Про доставку"
    ]
    for h in final_headers:
        if h not in df_all.columns:
            df_all[h] = np.nan
    df_final = df_all[final_headers].copy()
    df_final['Месяц'] = df_all['Месяц'].apply(get_month_name)
    df_final['Год'] = pd.to_numeric(df_final['Год'], errors="coerce").astype("Int16")
    df_final['Код продукта'] = pd.to_numeric(df_final['Код продукта'], errors="coerce").astype("Int64")
    df_final['Оценка'] = pd.to_numeric(df_final['Оценка'], errors="coerce").astype("Int8")
    for c in ["Хорошая", "Нейтральная", "Плохая"]:
        df_final[c] = pd.to_numeric(df_final[c], errors='coerce').astype("Int8")
    df_final.replace("", np.nan, inplace=True)

    print('Рассчитываю NPS...')
    dt = pd.to_datetime(df_final['Месяц и год'], dayfirst=True, errors='coerce')
    df_final['Дата_dt'] = dt
    df_final['_YM'] = list(zip(df_final['Дата_dt'].dt.year, df_final['Дата_dt'].dt.month))
    last3 = sorted(df_final['_YM'].dropna().unique())[-3:]
    mask = df_final['_YM'].isin(last3)
    df_3mo = df_final[mask].copy()

    def get_month_rus(date):
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                  "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        if pd.isnull(date):
            return np.nan
        if isinstance(date, str):
            try:
                date = pd.to_datetime(date, dayfirst=True)
            except Exception:
                return np.nan
        return f"{months[date.month - 1]} {date.year}"

    df_3mo['Месяц-год'] = df_3mo['Дата_dt'].apply(get_month_rus)
    months_order = (
        df_3mo
        .dropna(subset=['Дата_dt'])
        .sort_values('Дата_dt')
        ['Месяц-год']
        .drop_duplicates()
        .tolist()
    )

    print('Формирую итоговый вид...')
    # КЛЮЧЕВОЕ ИЗМЕНЕНИЕ: добавляем 'СТМ' к индексам!
    index_cols = ["Код продукта", "Продукт", "Поставщик", "Основной менеджер", "СТМ"]
    df_3mo[index_cols] = df_3mo[index_cols].fillna('')
    pivot = pd.pivot_table(
        df_3mo,
        index=index_cols,
        columns="Месяц-год",
        values=["Хорошая", "Нейтральная", "Плохая", "Результат"],
        aggfunc={"Хорошая": "sum", "Нейтральная": "sum", "Плохая": "sum", "Результат": "count"},
        fill_value=0,
        observed=True
    )
    pivot = pivot.rename(columns={"Результат": "Всего"}, level=0)
    pivot = pivot.sort_index(axis=1, level=1)
    pivot.columns = [' '.join(col).strip() for col in pivot.columns.values]
    pivot.reset_index(inplace=True)
    nps_cols, nps_count_cols = [], []
    for month in months_order:
        good = pivot.get(f"Хорошая {month}", 0)
        bad = pivot.get(f"Плохая {month}", 0)
        total = pivot.get(f"Всего {month}", 0)
        nps_col = f"NPS {month}"
        count_col = f"Количество отзывов {month}"
        with np.errstate(divide='ignore', invalid='ignore'):
            nps = np.where(total > 0, (good - bad) / total, np.nan)
        pivot[nps_col] = nps
        pivot[count_col] = total
        nps_cols.append(nps_col)
        nps_count_cols.append(count_col)
    good_total = pivot[[f"Хорошая {m}" for m in months_order]].sum(axis=1)
    bad_total = pivot[[f"Плохая {m}" for m in months_order]].sum(axis=1)
    total_total = pivot[[f"Всего {m}" for m in months_order]].sum(axis=1)
    pivot["NPS Общий итог"] = np.where(total_total > 0, (good_total - bad_total) / total_total, np.nan)
    pivot["Количество отзывов Общий итог"] = total_total

    print('Полировка...')
    pivot = pivot.rename(columns={
        "Код (доп.)": "Код продукта",
        "Продукт": "Номенклатура",
        "Основной менеджер": "Менеджер",
    })
    base_cols = [c for c in ["Код продукта", "Номенклатура", "Поставщик", "Менеджер", "СТМ"] if c in pivot.columns]

    bymonth_blocks = []
    for month in months_order:
        for met in ["Хорошая", "Нейтральная", "Плохая", "Всего"]:
            cname = f"{met} {month}"
            if cname in pivot.columns:
                bymonth_blocks.append(cname)
    nps_month_blocks = [f"NPS {month}" for month in months_order]
    counts_month_blocks = [f"Количество отзывов {month}" for month in months_order]
    total_blocks = [c for c in ["NPS Общий итог", "Количество отзывов Общий итог"] if c in pivot.columns]
    rest_cols = [c for c in pivot.columns if
                 c not in (base_cols + bymonth_blocks + nps_month_blocks + counts_month_blocks + total_blocks)]
    final_cols = base_cols + bymonth_blocks + nps_month_blocks + total_blocks + rest_cols
    pivot = pivot[final_cols]

    pivot = add_totals_row(pivot, base_cols, months_order)

    df_3mo.drop(columns=['Дата_dt', '_YM', 'Месяц-год'], inplace=True, errors='ignore')
    print('Сохраняю промежуточный вариант...')
    t9 = time.perf_counter()
    with pd.ExcelWriter(save_path) as writer:
        general_sheets_data = split_dataframe_to_sheets(df_3mo)
        for sheet_name, sheet_data in general_sheets_data.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
        pivot.to_excel(writer, sheet_name="по кодам", index=False)
    print(f'Готово! Сводная NPS сохранена: {save_path}')
    print(f'Время: {time.perf_counter() - t9:.2f} сек.')
    print(f'\nОбщее время: {time.perf_counter() - timer0:.2f} сек.')
    npsview_one(save_path, print)

if __name__ == "__main__":
    main()