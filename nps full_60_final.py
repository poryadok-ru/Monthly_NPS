import os
import pandas as pd
import numpy as np
import time
from datetime import datetime, timedelta
import math
import traceback

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font


def split_dataframe_to_sheets(df, max_rows=1000000):
    sheets_data = {}
    total_parts = math.ceil(len(df) / max_rows)
    for part in range(total_parts):
        start_idx = part * max_rows
        end_idx = min((part + 1) * max_rows, len(df))
        sheet_name = f"общее_часть{part + 1}" if part > 0 else "общее"
        sheets_data[sheet_name] = df.iloc[start_idx:end_idx].copy()
    return sheets_data


def get_month_name(month_number):
    months = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    try:
        return months[int(month_number) - 1] if not pd.isnull(month_number) else np.nan
    except Exception:
        return np.nan


def add_totals_row(pivot, base_cols, years_order):
    # Создаем словарь для итоговой строки
    itog_row = {}

    # Заполняем базовые колонки
    for bc in base_cols:
        itog_row[bc] = 'ИТОГО'

    # Собираем все числовые колонки, которые нужно просуммировать
    numeric_cols_to_sum = []

    # Колонки по годам: Хорошая, Плохая, Нейтральная, Количество отзывов, Всего
    for year in years_order:
        for metric in ["Хорошая", "Плохая", "Нейтральная", "Количество отзывов", "Всего"]:
            col_name = f"{metric} {year}"
            if col_name in pivot.columns:
                numeric_cols_to_sum.append(col_name)

    # Общие итоговые колонки
    if 'Количество отзывов Общий итог' in pivot.columns:
        numeric_cols_to_sum.append('Количество отзывов Общий итог')

    # Суммируем все числовые колонки
    for col in numeric_cols_to_sum:
        try:
            # Преобразуем в числовой формат и суммируем
            itog_row[col] = pd.to_numeric(pivot[col], errors='coerce').sum()
        except:
            itog_row[col] = np.nan

    # Рассчитываем NPS для каждого года
    for year in years_order:
        good_col = f"Хорошая {year}"
        bad_col = f"Плохая {year}"
        total_col = f"Всего {year}"
        nps_col = f"NPS {year}"

        if all(col in pivot.columns for col in [good_col, bad_col, total_col]):
            good_sum = pd.to_numeric(pivot[good_col], errors='coerce').sum()
            bad_sum = pd.to_numeric(pivot[bad_col], errors='coerce').sum()
            total_sum = pd.to_numeric(pivot[total_col], errors='coerce').sum()

            if total_sum > 0:
                itog_row[nps_col] = (good_sum - bad_sum) / total_sum
            else:
                itog_row[nps_col] = np.nan
        elif nps_col in pivot.columns:
            # Если NPS колонка есть, но нет исходных данных, оставляем NaN
            itog_row[nps_col] = np.nan

    # Рассчитываем общий NPS
    if 'NPS Общий итог' in pivot.columns:
        # Находим все колонки с данными по годам
        all_good_cols = [f"Хорошая {y}" for y in years_order if f"Хорошая {y}" in pivot.columns]
        all_bad_cols = [f"Плохая {y}" for y in years_order if f"Плохая {y}" in pivot.columns]
        all_total_cols = [f"Всего {y}" for y in years_order if f"Всего {y}" in pivot.columns]

        if all_good_cols and all_bad_cols and all_total_cols:
            sum_good = sum(pd.to_numeric(pivot[col], errors='coerce').sum() for col in all_good_cols)
            sum_bad = sum(pd.to_numeric(pivot[col], errors='coerce').sum() for col in all_bad_cols)
            sum_total = sum(pd.to_numeric(pivot[col], errors='coerce').sum() for col in all_total_cols)

            if sum_total > 0:
                itog_row['NPS Общий итог'] = (sum_good - sum_bad) / sum_total
            else:
                itog_row['NPS Общий итог'] = np.nan

    # Для остальных колонок ставим NaN
    for col in pivot.columns:
        if col not in itog_row:
            itog_row[col] = np.nan

    # Особое поле СТМ
    if 'СТМ' in pivot.columns:
        itog_row['СТМ'] = np.nan

    # Создаем итоговую строку и добавляем в DataFrame
    return pd.concat([pivot, pd.DataFrame([itog_row])], ignore_index=True)


# --- ЭТА ФУНКЦИЯ ДЕЛАЕТ ФИНАЛЬНОЕ ОФОРМЛЕНИЕ (цвет, группы и т.д.) ---
def npsview_format_file(excel_file):
    today = datetime.now()
    last_month = today.replace(day=1) - timedelta(days=1)
    try:
        xl_file = pd.ExcelFile(excel_file)
        sheet_names = xl_file.sheet_names
        general_sheets = [name for name in sheet_names if name.startswith('общее')]
        products_sheet = next((name for name in sheet_names if name == 'по кодам'), None)
        if not general_sheets or not products_sheet:
            raise KeyError("Не найдены необходимые листы")

        dfs_general = [pd.read_excel(excel_file, sheet_name=sheet) for sheet in general_sheets]
        df_comments = pd.concat(dfs_general, ignore_index=True)
        df_products = pd.read_excel(excel_file, sheet_name=products_sheet)

        if 'СТМ' not in df_products.columns:
            df_products['СТМ'] = np.nan
        stm_idx = df_products.columns.get_loc("СТМ")
        df_comments = df_comments[df_comments["Комментарий"].notna()]
        df_comments = df_comments[df_comments["Комментарий"].str.strip() != "-"]

        def safe_to_datetime(x):
            try:
                if pd.isna(x):
                    return pd.NaT
                if isinstance(x, (int, float)):
                    return pd.to_datetime("1899-12-30") + pd.to_timedelta(int(x), "D")
                return pd.to_datetime(x, errors="coerce", dayfirst=True)
            except Exception:
                return pd.NaT

        df_comments["Месяц и год"] = df_comments["Месяц и год"].apply(safe_to_datetime)
        df_comments = df_comments[df_comments["Месяц и год"].notna()]

        code_col = None
        for c in ["Код продукта", "Код (доп.)"]:
            if c in df_comments.columns and c in df_products.columns:
                code_col = c
                break
        if code_col is None:
            code_col = df_products.columns[0]

        product_comments = df_comments.groupby(code_col).apply(
            lambda x: list(zip(x["Комментарий"], x["Месяц и год"]))
        ).to_dict()

        rows_to_write = []
        is_code_row = []
        is_yellow_row = []
        for idx, row in df_products.iterrows():
            rows_to_write.append(list(row) + [None])
            is_code_row.append(True)
            is_yellow_row.append(False)
            code = row[code_col] if code_col in row else row.iloc[0]
            supplier = row.iloc[2] if len(row) > 2 else None
            manager = row.iloc[3] if len(row) > 3 else None
            has_yellow_comment = False
            comments_data = product_comments.get(code, [])
            comment_rows = []
            yellow_flags = []
            for comment, cdate in comments_data:
                is_yellow = (
                        pd.notnull(cdate) and
                        isinstance(cdate, pd.Timestamp) and
                        cdate.year == last_month.year and
                        cdate.month == last_month.month
                )
                comment_row = [None] * len(row)
                comment_row[1] = comment
                if len(row) > 2:
                    comment_row[2] = supplier
                if len(row) > 3:
                    comment_row[3] = manager
                comment_row[stm_idx] = row['СТМ'] if 'СТМ' in row else np.nan
                comment_row.append(is_yellow)
                comment_rows.append(comment_row)
                yellow_flags.append(is_yellow)
                if is_yellow:
                    has_yellow_comment = True
            rows_to_write.extend(comment_rows)
            is_code_row.extend([False] * len(comment_rows))
            is_yellow_row.extend(yellow_flags)
            if has_yellow_comment:
                is_yellow_row[-(len(comment_rows) + 1)] = True
        out_df = pd.DataFrame(rows_to_write)
        out_df.columns = list(df_products.columns) + ["_yellow"]

        # Сохраняем только структуру "по кодам" (для форматирования)
        with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            out_df.drop(columns="_yellow").to_excel(writer, sheet_name="по кодам", index=False)

        wb = load_workbook(excel_file)
        try:
            if "по кодам" in wb.sheetnames:
                sh = wb["по кодам"]
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                wrap_align = Alignment(wrap_text=True, vertical='center')
                header_font = Font(bold=True)
                is_code_row2 = [0] + is_code_row
                is_yellow_row2 = [False] + is_yellow_row
                max_row = sh.max_row
                max_col = sh.max_column
                for col in range(1, max_col + 1):
                    cell = sh.cell(row=1, column=col)
                    cell.alignment = wrap_align
                    cell.font = header_font
                    header_value = cell.value
                    if header_value:
                        header_str = str(header_value)
                        if "NPS" in header_str and "Общий" in header_str:
                            cell.fill = green_fill
                        elif "NPS" in header_str:
                            cell.fill = blue_fill
                nps_cols = [col for col in range(1, max_col + 1)
                            if sh.cell(row=1, column=col).value and str(sh.cell(row=1, column=col).value).startswith(
                        "NPS ")]
                for col in nps_cols:
                    for row in range(2, max_row + 1):
                        try:
                            sh.cell(row=row, column=col).number_format = '0%'
                        except Exception as e:
                            print(f"DIAG Ошибка при форматировании процентов: строка {row}, колонка {col}: {e}")
                painted = 0
                for i in range(2, max_row + 1):
                    if is_yellow_row2[i - 1]:
                        try:
                            for col in range(1, max_col + 1):
                                sh.cell(row=i, column=col).fill = yellow_fill
                            painted += 1
                        except Exception as e:
                            print(f"DIAG Ошибка при покраске строки {i}: {e}")
                            print(traceback.format_exc())
                i = 2
                while i <= max_row:
                    if is_code_row2[i - 1]:
                        group_start = i + 1
                        group_end = group_start
                        while group_end <= max_row and not is_code_row2[group_end - 1]:
                            group_end += 1
                        if group_end > group_start:
                            try:
                                sh.row_dimensions.group(start=group_start, end=group_end - 1, hidden=False)
                            except Exception as e:
                                print(f"DIAG Ошибка при группировке строк {group_start}-{group_end - 1}: {e}")
                                print(traceback.format_exc())
                        i = group_end
                    else:
                        i += 1
                for col in range(1, max_col + 1):
                    try:
                        cell = sh.cell(row=max_row, column=col)
                        cell.font = Font(bold=True)
                    except Exception as e:
                        print(f"DIAG Ошибка при выделении последней строки: {e}")
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith('общее'):
                    sh = wb[sheet_name]
                    max_col = sh.max_column
                    for col in range(1, max_col + 1):
                        try:
                            cell = sh.cell(row=1, column=col)
                            cell.alignment = Alignment(wrap_text=True, vertical='center')
                            cell.font = Font(bold=True)
                        except Exception as e:
                            print(f"DIAG Ошибка при форматировании заголовка в листе {sheet_name}: {e}")
            wb.save(excel_file)
            print(f"Форматирование и группировки добавлены: '{excel_file}'")
        except Exception as e:
            print(f"DIAG: Ошибка на этапе openpyxl: {e}")
            print(traceback.format_exc())
    except Exception as e:
        print(f"Произошла ошибка при оформлении: {e}")
        print(traceback.format_exc())


# --- Эта функция УДАЛЯЕТ группы с NPS > 0.6 на всех листах и пересчитывает ИТОГО ---
def filter_groups_by_nps_inplace(upd_path):
    import warnings
    warnings.filterwarnings("ignore")
    xls = pd.ExcelFile(upd_path, engine="openpyxl")
    po_kodam = pd.read_excel(xls, sheet_name="по кодам", dtype=object)
    all_sheet_names = xls.sheet_names
    obschee_sheets = [s for s in all_sheet_names if s.lower().startswith('общее')]
    obschee_dict = {s: pd.read_excel(upd_path, sheet_name=s, dtype=object) for s in obschee_sheets}
    col_code = "Код продукта"
    col_nps = "NPS Общий итог"
    mask_code = po_kodam[col_code].notna() & (~po_kodam[col_code].astype(str).str.upper().eq("ИТОГО"))
    nps = pd.to_numeric(po_kodam[col_nps], errors='coerce')
    codes_to_del = set(po_kodam.loc[mask_code & (nps > 0.6), col_code].astype(str))
    # Удаляем группы строк на листе по кодам (код и все комментарии ниже)
    rows_to_drop = set()
    i = 0
    while i < len(po_kodam):
        row = po_kodam.iloc[i]
        kod = str(row[col_code]) if pd.notnull(row[col_code]) else None
        if kod and kod not in ["ИТОГО", "nan"] and kod in codes_to_del:
            rows_to_drop.add(i)
            i += 1
            while i < len(po_kodam) and pd.isnull(po_kodam.iloc[i][col_code]):
                rows_to_drop.add(i)
                i += 1
        else:
            i += 1
    po_kodam_wo_groups = po_kodam.loc[~po_kodam.index.isin(rows_to_drop)].reset_index(drop=True)
    itogo_mask = po_kodam_wo_groups[col_code].astype(str).str.upper().eq("ИТОГО")
    po_kodam_nobottom = po_kodam_wo_groups[~itogo_mask]

    def calc_itogo(df, id_col="Код продукта"):
        code_rows = df[df[id_col].notna() & ~df[id_col].astype(str).str.upper().eq("ИТОГО")]
        itogo = {}

        # Базовые колонки
        for col in df.columns:
            if col == id_col:
                itogo[col] = "ИТОГО"
            elif np.issubdtype(df[col].dtype, np.number) or df[col].dtype == 'object':
                # Пробуем преобразовать в число и суммировать
                numeric_values = pd.to_numeric(code_rows[col], errors='coerce')
                if not numeric_values.isna().all():  # Если есть числовые значения
                    itogo[col] = numeric_values.sum()
                else:
                    itogo[col] = np.nan
            else:
                itogo[col] = np.nan

        # Рассчитываем NPS для каждого года
        all_cols = [str(c) for c in df.columns]
        years = []
        for col in all_cols:
            if col.startswith("Хорошая "):
                year = col.replace("Хорошая ", "")
                years.append(year)

        for year in set(years):
            nps_col = f"NPS {year}"
            good_col = f"Хорошая {year}"
            bad_col = f"Плохая {year}"
            total_col = f"Всего {year}"

            if all(c in df.columns for c in [good_col, bad_col, total_col]):
                good_sum = pd.to_numeric(code_rows[good_col], errors='coerce').sum()
                bad_sum = pd.to_numeric(code_rows[bad_col], errors='coerce').sum()
                total_sum = pd.to_numeric(code_rows[total_col], errors='coerce').sum()

                if total_sum > 0:
                    itogo[nps_col] = (good_sum - bad_sum) / total_sum
                else:
                    itogo[nps_col] = np.nan

        # Рассчитываем общий NPS
        if "NPS Общий итог" in df.columns:
            # Находим все колонки с хорошими, плохими и всего
            good_cols = [c for c in df.columns if str(c).startswith("Хорошая ")]
            bad_cols = [c for c in df.columns if str(c).startswith("Плохая ")]
            total_cols = [c for c in df.columns if str(c).startswith("Всего ")]

            if good_cols and bad_cols and total_cols:
                gsum = sum(pd.to_numeric(code_rows[col], errors='coerce').sum() for col in good_cols)
                bsum = sum(pd.to_numeric(code_rows[col], errors='coerce').sum() for col in bad_cols)
                tsum = sum(pd.to_numeric(code_rows[col], errors='coerce').sum() for col in total_cols)

                if tsum > 0:
                    itogo["NPS Общий итог"] = (gsum - bsum) / tsum
                else:
                    itogo["NPS Общий итог"] = np.nan

        # Особое поле СТМ
        if 'СТМ' in df.columns:
            itogo['СТМ'] = np.nan

        return pd.DataFrame([itogo])

    itogo_row = calc_itogo(po_kodam_nobottom)
    po_kodam_final = pd.concat([po_kodam_nobottom, itogo_row], ignore_index=True)
    # Удалить эти коды из всех "общее..." листов
    obschee_dict2 = {}
    for sheetname, df in obschee_dict.items():
        if col_code in df.columns:
            obschee_dict2[sheetname] = df[~df[col_code].astype(str).isin(codes_to_del)].copy()
        else:
            obschee_dict2[sheetname] = df.copy()
    # Сохранить все изменённые листы обратно
    with pd.ExcelWriter(upd_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        po_kodam_final.to_excel(writer, sheet_name="по кодам", index=False)
        for sn, df in obschee_dict2.items():
            df.to_excel(writer, sheet_name=sn, index=False)
    print(f"Удаление групп с высоким NPS выполнено (удалено кодов: {len(codes_to_del)}), итоговая строка пересчитана!")


def main():
    timer0 = time.perf_counter()
    folder = 'Исходники WB и Ozon'
    card_path = os.path.join(folder, 'Карточка товара.xlsx')
    save_path = os.path.join(folder, 'WB&OZON NPS full_60%.xlsx')
    print('Старт. Проверка условий.')
    print('Читаю файлы...')
    t1 = time.perf_counter()
    files = [f for f in os.listdir(folder)
             if (f.endswith('.xlsx') or f.endswith('.csv')) and not ('nps' in f.lower() or 'карточк' in f.lower())]
    dataframes = []
    for f in files:
        path = os.path.join(folder, f)
        try:
            if f.endswith('.csv'):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, dtype=str)
            dataframes.append(df)
            print(f'  Прочитан: {f} ({len(df)} строк)')
        except Exception as e:
            print(f'Ошибка в файле {f}: {e}')
    if len(dataframes) == 0:
        print('Нет исходных данных!')
        return
    df_all = pd.concat(dataframes, ignore_index=True)
    print(f'Считано файлов: {len(dataframes)}, итоговых строк: {df_all.shape[0]}')
    print(f'Время чтения файлов: {time.perf_counter() - t1:.2f} сек.')
    print('Собираю файл...')
    rename_map = {
        "COMMENT": "Комментарий",
        "CREATED_AT": "Месяц и год",
        "PRODUCT.CODE": "Код продукта",
        "RATING": "Оценка",
        "SOURCE": "Источник",
        "IS_ABOUT_DELIVERY": "Про доставку"
    }
    df_all = df_all.rename(columns={col: rename_map[col] for col in df_all.columns if col in rename_map})
    print('Преобразовываю даты...')
    df_all['date_clean'] = df_all['Месяц и год'].fillna('').str.split('@').str[0].str.strip()
    dt = pd.to_datetime(df_all['date_clean'], format='%b %d, %Y', errors='coerce')
    not_parsed = dt.isna()
    if not_parsed.any():
        dt2 = pd.to_datetime(df_all.loc[not_parsed, 'date_clean'], dayfirst=True, errors='coerce')
        dt.loc[not_parsed] = dt2
    df_all['Дата_dt'] = dt
    df_all['Месяц и год'] = df_all['Дата_dt'].dt.date
    df_all['Год'] = df_all['Дата_dt'].dt.year
    df_all['Месяц'] = df_all['Дата_dt'].dt.month
    print('Работаю с оценками...')
    df_all["Результат"] = (
        df_all["Оценка"].map({'5': "Хорошая", '4': "Нейтральная", '3': "Плохая", '2': "Плохая", '1': "Плохая"})
    )
    df_all["Хорошая"] = (df_all["Результат"] == "Хорошая").astype("Int8")
    df_all["Нейтральная"] = (df_all["Результат"] == "Нейтральная").astype("Int8")
    df_all["Плохая"] = (df_all["Результат"] == "Плохая").astype("Int8")
    print('Заполяняю данные из Карточки товаров...')
    t5 = time.perf_counter()
    extra_cols = ["Продукт", "Основной менеджер", "СегментСтелажногоХранения", "Подгруппа 1", "Поставщик", "СТМ"]
    if os.path.exists(card_path):
        card_df = pd.read_excel(card_path, dtype=str)
        card_df = card_df.rename(columns=lambda x: x.strip())
        for col in extra_cols:
            if col not in card_df.columns:
                card_df[col] = np.nan
        card_df.set_index("Код (доп.)", inplace=True)
        df_all = df_all.join(card_df[extra_cols], on="Код продукта", how="left")
    else:
        for c in extra_cols:
            df_all[c] = np.nan
    print('Играю в нарды...')
    final_headers = [
        "Комментарий", "Месяц и год", "Год", "Месяц", "Код продукта", "Продукт", "Основной менеджер",
        "Оценка", "Результат", "Хорошая", "Нейтральная", "Плохая", "Источник",
        "СегментСтелажногоХранения", "Подгруппа 1", "Поставщик", "СТМ", "Про доставку"
    ]
    for h in final_headers:
        if h not in df_all.columns:
            df_all[h] = np.nan
    df_final = df_all[final_headers].copy()
    df_final['Месяц'] = df_all['Месяц'].apply(get_month_name)
    df_final['Год'] = pd.to_numeric(df_final['Год'], errors="coerce").astype("Int16")
    df_final['Код продукта'] = pd.to_numeric(df_final['Код продукта'], errors="coerce").astype("Int64")
    df_final['Оценка'] = pd.to_numeric(df_final['Оценка'], errors="coerce").astype("Int8")
    for c in ["Хорошая", "Нейтральная", "Плохая"]:
        df_final[c] = pd.to_numeric(df_final[c], errors='coerce').astype("Int8")
    df_final.replace("", np.nan, inplace=True)
    print('Рассчитываю NPS...')
    dt = pd.to_datetime(df_final['Месяц и год'], dayfirst=True, errors='coerce')
    df_final['Дата_dt'] = dt
    df_final['_year'] = df_final['Дата_dt'].dt.year
    df_by_year = df_final.copy()
    df_by_year['Год_cat'] = df_by_year['Год'].astype('Int64').astype(str)
    years_order = (
        df_by_year
        .dropna(subset=['Год_cat'])
        .sort_values('Год')
        ['Год_cat']
        .drop_duplicates()
        .tolist()
    )
    print('Формирую итоговый вид...')
    index_cols = ["Код продукта", "Продукт", "Поставщик", "Основной менеджер", "СТМ"]
    df_by_year[index_cols] = df_by_year[index_cols].fillna('')
    pivot = pd.pivot_table(
        df_by_year,
        index=index_cols,
        columns="Год_cat",
        values=["Хорошая", "Нейтральная", "Плохая", "Результат"],
        aggfunc={"Хорошая": "sum", "Нейтральная": "sum", "Плохая": "sum", "Результат": "count"},
        fill_value=0,
        observed=True
    )
    pivot = pivot.rename(columns={"Результат": "Всего"}, level=0)
    pivot = pivot.sort_index(axis=1, level=1)
    pivot.columns = [' '.join(col).strip() for col in pivot.columns.values]
    pivot.reset_index(inplace=True)
    nps_cols, nps_count_cols = [], []
    for year in years_order:
        good = pivot.get(f"Хорошая {year}", 0)
        bad = pivot.get(f"Плохая {year}", 0)
        total = pivot.get(f"Всего {year}", 0)
        nps_col = f"NPS {year}"
        count_col = f"Количество отзывов {year}"
        with np.errstate(divide='ignore', invalid='ignore'):
            nps = np.where(total > 0, (good - bad) / total, np.nan)
        pivot[nps_col] = nps
        pivot[count_col] = total
        nps_cols.append(nps_col)
        nps_count_cols.append(count_col)
    good_total = pivot[[f"Хорошая {y}" for y in years_order]].sum(axis=1)
    bad_total = pivot[[f"Плохая {y}" for y in years_order]].sum(axis=1)
    total_total = pivot[[f"Всего {y}" for y in years_order]].sum(axis=1)
    pivot["NPS Общий итог"] = np.where(total_total > 0, (good_total - bad_total) / total_total, np.nan)
    pivot["Количество отзывов Общий итог"] = total_total
    print('Полировка...')
    pivot = pivot.rename(columns={
        "Код (доп.)": "Код продукта",
        "Продукт": "Номенклатура",
        "Основной менеджер": "Менеджер",
    })
    base_cols = [c for c in ["Код продукта", "Номенклатура", "Поставщик", "Менеджер", "СТМ"] if c in pivot.columns]
    byyear_blocks = []
    for year in years_order:
        for met in ["Хорошая", "Нейтральная", "Плохая", "Всего"]:
            cname = f"{met} {year}"
            if cname in pivot.columns:
                byyear_blocks.append(cname)
    nps_year_blocks = [f"NPS {year}" for year in years_order]
    counts_year_blocks = [f"Количество отзывов {year}" for year in years_order]
    total_blocks = [c for c in ["NPS Общий итог", "Количество отзывов Общий итог"] if c in pivot.columns]
    rest_cols = [c for c in pivot.columns if
                 c not in (base_cols + byyear_blocks + nps_year_blocks + counts_year_blocks + total_blocks)]
    final_cols = base_cols + byyear_blocks + nps_year_blocks + total_blocks + rest_cols
    pivot = pivot[final_cols]
    pivot = add_totals_row(pivot, base_cols, years_order)
    df_by_year.drop(columns=['Дата_dt', '_year', 'Год_cat'], inplace=True, errors='ignore')
    print('Сохраняю промежуточный вариант...')
    t9 = time.perf_counter()
    with pd.ExcelWriter(save_path) as writer:
        general_sheets_data = split_dataframe_to_sheets(df_by_year)
        for sheet_name, sheet_data in general_sheets_data.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
        pivot.to_excel(writer, sheet_name="по кодам", index=False)
    print(f'Готово! Сводная NPS сохранена: {save_path}')
    print(f'Время: {time.perf_counter() - t9:.2f} сек.')
    print(f'\nОбщее время: {time.perf_counter() - timer0:.2f} сек.')
    # ==== Дальше: фильтр и только после него — оформление! ====
    upd_path = os.path.splitext(save_path)[0] + '_upd.xlsx'
    # Переносим файл для дальнейшей обработки
    import shutil
    shutil.copy(save_path, upd_path)
    filter_groups_by_nps_inplace(upd_path)
    npsview_format_file(upd_path)


if __name__ == "__main__":
    main()